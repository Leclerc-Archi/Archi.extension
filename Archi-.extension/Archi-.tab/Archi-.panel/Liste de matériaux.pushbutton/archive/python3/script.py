#! python3
# -*- coding: utf-8 -*-
import clr, sys, os

from Autodesk.Revit.UI import TaskDialog
import System.Windows.Forms as WinForms

# --- find the real button folder (works even if pyRevit runs from a temp copy) ---
try:
    from pyrevit import script as _script
    _BUNDLE_DIR = _script.get_bundle_dir()  # reliable bundle path
except Exception:
    # fallback to where this file lives
    _BUNDLE_DIR = os.path.dirname(__file__)

# --- add lib candidates to sys.path (bundle/lib and a couple of parents just in case) ---
def _add_libs():
    bases = [
        _BUNDLE_DIR,
        os.path.dirname(_BUNDLE_DIR),
        os.path.dirname(os.path.dirname(_BUNDLE_DIR)),
    ]
    added = []
    for base in bases:
        libp = os.path.join(base, 'lib')
        if os.path.isdir(libp) and libp not in sys.path:
            sys.path.insert(0, libp)
            added.append(libp)
            # allow dropping .whl/.zip into lib
            try:
                for name in os.listdir(libp):
                    if name.lower().endswith(('.whl', '.zip')):
                        p = os.path.join(libp, name)
                        if p not in sys.path:
                            sys.path.insert(0, p)
                            added.append(p)
            except Exception:
                pass
    return added

_added = _add_libs()

# --- import openpyxl from vendored lib ---
try:
    from openpyxl import load_workbook
except ImportError:
    msg = (
        u"Module 'openpyxl' introuvable.\n\n"
        u"Dossier du bouton (bundle):\n{}\n\n"
        u"Chemins ajoutés :\n{}\n\n"
        u"Vérifiez que l'un contient:\n  lib\\openpyxl\\__init__.py\n  lib\\et_xmlfile\\__init__.py\n"
        u"Sinon, copiez les dossiers 'openpyxl' et 'et_xmlfile' dans:\n  {}\\lib"
    ).format(_BUNDLE_DIR, "\n".join(_added or ["(aucun)"]), _BUNDLE_DIR)
    TaskDialog.Show('Erreur', msg)
    sys.exit(1)

# (keep your existing imports below)
from System import Array, String
from Autodesk.Revit.DB import *
from pyrevit import revit, script

# 1. document actif
doc = revit.doc
if not doc:
    TaskDialog.Show('Erreur', 'Aucun document Revit actif.')
    sys.exit()

# 2. sélection des nommenclatures dans Revit
nommenclatures = list(FilteredElementCollector(doc).OfClass(ViewSchedule))
noms_nomms = sorted([nm.Name for nm in nommenclatures])
if not noms_nomms:
    TaskDialog.Show('Erreur', 'Aucun type de pochage trouvé.')
    sys.exit()

idx_nomm = noms_nomms.index('LISTE DES MATÉRIAUX ET FINIS') if 'LISTE DES MATÉRIAUX ET FINIS' in noms_nomms else 0

# 3. données utilisateur (sans persistance)
def afficher_formulaire():
    posX = 10
    posY = 10
    ligneH = 30

    form = WinForms.Form()
    form.Text = 'Liste des matériaux'
    form.Width = 500
    form.Height = 450
    form.StartPosition = WinForms.FormStartPosition.CenterScreen

    # sélecteur de nommenclature (label)
    lblNomm = WinForms.Label()
    lblNomm.Text = 'Sélectionner la nommenclature :'
    lblNomm.Left = posX
    lblNomm.Top = posY
    lblNomm.Height = ligneH
    lblNomm.Width = 340
    form.Controls.Add(lblNomm)

    posY += 40
    comboNomm = WinForms.ComboBox()
    comboNomm.Left = posX
    comboNomm.Top = posY
    comboNomm.Height = ligneH
    comboNomm.Width = 460
    comboNomm.DropDownStyle = WinForms.ComboBoxStyle.DropDownList
    comboNomm.Items.AddRange(Array[String](noms_nomms))
    comboNomm.SelectedIndex = idx_nomm
    form.Controls.Add(comboNomm)

    # sélecteur de fichier Excel (label)
    posY += 60
    lblExcel = WinForms.Label()
    lblExcel.Text = 'Fichier excel :'
    lblExcel.Left = posX
    lblExcel.Top = posY
    lblExcel.Height = ligneH
    lblExcel.Width = 340
    form.Controls.Add(lblExcel)

    posY += 40
    txtPath = WinForms.TextBox()
    txtPath.Left = posX
    txtPath.Top = posY
    txtPath.Height = ligneH
    txtPath.Width = 460
    txtPath.ReadOnly = True
    txtPath.Text = ''
    form.Controls.Add(txtPath)

    # bouton 'Parcourir'
    posY += -40
    btnBrowse = WinForms.Button()
    btnBrowse.Text = u'Parcourir'
    btnBrowse.Left = posX + 360
    btnBrowse.Top = posY
    btnBrowse.Height = ligneH
    btnBrowse.Width = 100
    form.Controls.Add(btnBrowse)

    # sélecteur de feuille (label)
    posY += 100
    lblFeuille = WinForms.Label()
    lblFeuille.Text = u'Feuille :'
    lblFeuille.Left = posX
    lblFeuille.Top = posY
    lblFeuille.Height = ligneH
    lblFeuille.Width = 460
    form.Controls.Add(lblFeuille)

    posY += 40
    comboFeuille = WinForms.ComboBox()
    comboFeuille.Left = posX
    comboFeuille.Top = posY
    comboFeuille.Height = ligneH
    comboFeuille.Width = 460
    comboFeuille.DropDownStyle = WinForms.ComboBoxStyle.DropDownList
    form.Controls.Add(comboFeuille)

    # handlers
    def browse_click(sender, arg):
        dlg = WinForms.OpenFileDialog()
        dlg.Filter = 'Fichiers Excel (*.xlsx;*.xls)|*.xlsx;*.xls'
        if dlg.ShowDialog() != WinForms.DialogResult.OK:
            return
        txtPath.Text = dlg.FileName
        # lecture des noms de feuilles (openpyxl)
        noms_feuilles = []
        wb = None
        try:
            wb = load_workbook(txtPath.Text, data_only=True, read_only=True)
            noms_feuilles = wb.sheetnames
        except Exception as e:
            WinForms.MessageBox.Show('Erreur lecture Excel :\n{}'.format(str(e)))
        finally:
            try:
                if wb:
                    wb.close()
            except:
                pass
            comboFeuille.Items.Clear()

        if noms_feuilles:
            comboFeuille.Items.AddRange(Array[String](noms_feuilles))
            idx_feuille = noms_feuilles.index('PROJET') if 'PROJET' in noms_feuilles else 0
            comboFeuille.SelectedIndex = idx_feuille

    btnBrowse.Click += browse_click

    # bouton OK
    posY += 60
    btnOK = WinForms.Button()
    btnOK.Text = 'OK'
    btnOK.DialogResult = WinForms.DialogResult.OK
    btnOK.Left = posX + 185
    btnOK.Top = posY
    btnOK.Height = ligneH
    btnOK.Width = 80
    form.AcceptButton = btnOK
    form.Controls.Add(btnOK)

    # show dialog
    if form.ShowDialog() == WinForms.DialogResult.OK:
        return (comboNomm.Text, txtPath.Text, comboFeuille.Text)
    return None, None, None

# 4. récupération des données utilisateur
nom_nomm, path_fichier_excel, nom_feuille_excel = afficher_formulaire()
if not all([nom_nomm, path_fichier_excel, nom_feuille_excel]):
    sys.exit()

# 5. valeurs excel (openpyxl)
wb = load_workbook(path_fichier_excel, data_only=True, read_only=True)
feuille_excel = wb[nom_feuille_excel]

header_idx = 3  # (ligne 4, 0-based)
headers = next(feuille_excel.iter_rows(min_row=header_idx+1, max_row=header_idx+1, values_only=True))

num = headers.index('CODE')
descr = headers.index('DESCRIPTION')
rev = headers.index('REV.')

data = []

def valeur_cell(row_idx, col_idx):
    # conserver la sémantique 0-based
    return feuille_excel.cell(row=row_idx + 1, column=col_idx + 1).value

nrows = feuille_excel.max_row
for row in range(5, nrows):
    data.append([valeur_cell(row, num), valeur_cell(row, descr), valeur_cell(row, rev)])

wb.close()

# 6. nommenclature et écriture
ajour, erreur, tbshoot = [], [], []
nbLignes = len(data)

with revit.Transaction('écriture nommenclature'):
    for nm in nommenclatures:
        if str(nm.Name) == nom_nomm:
            elements = FilteredElementCollector(doc, nm.Id).ToElements()
            for e in elements:
                doc.Delete(e.Id)

            for i in range(nbLignes):
                table = nm.GetTableData()
                tableSection = table.GetSectionData(SectionType.Body)
                newe = tableSection.InsertRow(0)

            newElements = FilteredElementCollector(doc, nm.Id).ToElements()

            nb = 0
            for num, descr, rev in data:
                ligne = newElements[nb]
                nb = nb + 1
                try:
                    ligne.LookupParameter('Nom de la clé').Set(num or '')
                    ligne.LookupParameter('Commentaires').Set(descr or '')
                    ligne.LookupParameter('RÉVISION').Set(rev or '')
                    if num and num[-2:] == "xx":
                        ligne.LookupParameter('SECTION').Set("xx")
                    if num is not None:
                        ajour.append('  {0} - {1}\n'.format(num, descr))
                except:
                    if num is None:
                        pass
                    else:
                        erreur.append('  {0} - {1}\n'.format(num, descr))

TaskDialog.Show('Tâche complétée', 'Liste des matériaux à jour')